# Propios
from conn import Conexion
# Externos
from datetime import datetime, timedelta
import pandas as pd
import warnings
import numpy as np
from openpyxl import load_workbook
from time import time, gmtime
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)

class Sap:

    """Querys para traer la data de las txn de SAP"""

    def fbl3n(self, amb, fh_consulta):
        
        #ctas = Parametros.codigo_cuentas()
        fh_consulta_sap = fh_consulta.replace("/","")

        try:
            conn = Conexion()
            conn = conn.connection_bd(amb)
            cursor = conn.cursor()
            cursor.execute("SET SCHEMA SAPABAP1")
            print(f'Conexion a base de datos fbl3n ok')
        except Exception as e:
            print(f'Error en conexion a base de datos: {e}')

        q_3n = f"""

        WITH T1 AS (
        -- # FBL3N # SIN 121050/121080/711201--
        SELECT
        T.HKONT AS LIBRO_MAYOR,
        CAST(T2.BUDAT AS DATE) AS FH_CONTAB,
        T.BELNR AS NRO_DOC,
        T2.BLART AS CLASE,
        CAST(T2.BLDAT AS DATE) AS FH_DOC,
        CASE
        WHEN T.SHKZG = 'H' THEN -(T.DMBTR)
        ELSE T.DMBTR
        END AS IMPORTE_EN_ML,
        T.MWSKZ AS II,
        '--------->' AS AUXILIARES_3N,
        CAST(T.ZFBDT AS DATE) AS FH_BASE,
        T.AUGDT AS COMPENSACION,
        T.KUNNR AS SOLICITANTE,
        T.ZTERM AS COND_PAGO,
        T.ZBD1T AS PLAZO_PAGO

        FROM
        SAPABAP1.BSEG T
        LEFT JOIN SAPABAP1.BKPF T2 ON
        T.BELNR = T2.BELNR
        LEFT  JOIN SAPABAP1.KNB1 T3 ON
        T.KUNNR = T3.KUNNR
        LEFT JOIN SAPABAP1.KNA1 T4 ON
        T.KUNNR = T4.KUNNR
        WHERE
        T.BUKRS = 'SC10'
        AND T.HKONT IN (
        '0000121010', -- OK
        '0000121020', -- OK
        '0000121030', -- OK
        '0000112010', -- OK
        '0000123010', -- OK
        '0000712200', -- OK
        '0000712400', -- OK
        '0000711202') -- OK
        -- ================================
        -- Si es 0000711201 ->  KOART = 'K'
        -- 0000121050/0000121080 -> KOART = 'S'
        -- 0000122010 Gestion judicial se trata aparte
        -- ===============================
        AND T.KOART = 'D'
        AND (T.AUGDT = '00000000' OR T.AUGDT > {fh_consulta_sap}) AND T2.BUDAT <= {fh_consulta_sap}
        ORDER BY
        T2.BELNR ASC
        ),

        -- # FBL3N  para 711201 # --
        T2 AS (
        SELECT
        T.HKONT AS LIBRO_MAYOR,
        CAST(T2.BUDAT AS DATE) AS FH_CONTAB,
        T.BELNR AS NRO_DOC,
        T2.BLART AS CLASE,
        CAST(T2.BLDAT AS DATE) AS FH_DOC,
        CASE
        WHEN T.SHKZG = 'H' THEN -(T.DMBTR)
        ELSE T.DMBTR
        END AS IMPORTE_EN_ML,
        T.MWSKZ AS II,
        '--------->' AS AUXILIARES_3N,
        CAST(T.ZFBDT AS DATE) AS FH_BASE,
        T.AUGDT AS COMPENSACION,
        T.KUNNR AS SOLICITANTE,
        T.ZTERM AS COND_PAGO,
        T.ZBD1T AS PLAZO_PAGO
        FROM
        SAPABAP1.BSEG T
        LEFT JOIN SAPABAP1.BKPF T2 ON
        T.BELNR = T2.BELNR
        LEFT  JOIN SAPABAP1.KNB1 T3 ON
        T.KUNNR = T3.KUNNR
        LEFT JOIN SAPABAP1.KNA1 T4 ON
        T.KUNNR = T4.KUNNR
        WHERE
        T.BUKRS = 'SC10'
        AND T.HKONT IN (
        '0000711201') -- OK KOART = 'K'
        AND  T.KOART = 'K'
        AND (T.AUGDT = '00000000' OR T.AUGDT > {fh_consulta_sap}) AND T2.BUDAT <= {fh_consulta_sap}
        ORDER BY
        T2.BELNR ASC
        ),

        -- # FBL3N  para 121050/121080 # --
        T3 AS (
        SELECT
        T.HKONT AS LIBRO_MAYOR,
        CAST(T2.BUDAT AS DATE) AS FH_CONTAB,
        T.BELNR AS NRO_DOC,
        T2.BLART AS CLASE,
        CAST(T2.BLDAT AS DATE) AS FH_DOC,
        CASE
        WHEN T.SHKZG = 'H' THEN -(T.DMBTR)
        ELSE T.DMBTR
        END AS IMPORTE_EN_ML,
        T.MWSKZ AS II,
        '--------->' AS AUXILIARES_3N,
        CAST(T.ZFBDT AS DATE) AS FH_BASE,
        T.AUGDT AS COMPENSACION,
        T.KUNNR AS SOLICITANTE,
        T.ZTERM AS COND_PAGO,
        T.ZBD1T AS PLAZO_PAGO
        FROM
        SAPABAP1.BSEG T
        LEFT JOIN SAPABAP1.BKPF T2 ON
        T.BELNR = T2.BELNR
        LEFT  JOIN SAPABAP1.KNB1 T3 ON
        T.KUNNR = T3.KUNNR
        LEFT JOIN SAPABAP1.KNA1 T4 ON
        T.KUNNR = T4.KUNNR
        WHERE
        T.BUKRS = 'SC10'
        AND T.HKONT IN (
        '0000121050', -- OK KOART = 'S'
        '0000121080') -- OK KOART = 'S'
        AND  T.KOART = 'S'
        AND (T.AUGDT = '00000000' OR T.AUGDT > {fh_consulta_sap}) AND T2.BUDAT <= {fh_consulta_sap}
        ORDER BY
        T2.BELNR ASC
        )

        --------------------------------------------------------------
        SELECT * FROM T1 UNION ALL
        SELECT * FROM T2 UNION ALL
        SELECT * FROM T3
        """ 

        cursor.execute(q_3n)
        df_fbl3n = pd.read_sql_query(q_3n, conn)
        print("\t"+f"FBL3N: Se enconraron {df_fbl3n.shape[0]} partidas a fecha {fh_consulta}")
        df_fbl3n.fillna(0,inplace=True)


        return df_fbl3n


    def fbl5n(self, amb, fh_consulta):
        
        
        fh_consulta_sap = fh_consulta.replace("/","")

        try:
            conn = Conexion()
            conn = conn.connection_bd(amb)
            cursor = conn.cursor()
            cursor.execute("SET SCHEMA SAPABAP1")
            print(f'Conexion a base de datos fbl5n ok')
        except Exception as e:
            print(f'Error en conexion a base de datos: {e} ')

        q_5n = f"""        
        -- # FBL5N # -- 
        SELECT
        T.KUNNR AS CUENTA,
        T.HKONT AS LIBRO_MAYOR,
        T4.NAME1 AS CLIENTE,
        T2.BLART AS CL,
        T.BSCHL AS CT,
        T.ZLSCH AS VP,
        T.UMSKZ AS CME,
        T.BELNR AS NRO_DOC,
        T.REBZG AS REF_FACT,
        T2.XBLNR AS REFERENCIA,
        T.XREF3 AS CLAVE_REFERENCIA,
        T.ZUONR AS ASIGNACION,
        CAST(T2.BLDAT AS DATE) AS FH_DOC,
        CAST(T.ZFBDT AS DATE) AS FH_BASE,
        'FH_VTO_NETO' AS VTO_NETO, 
        CASE WHEN T.SHKZG = 'H' THEN -(T.DMBTR) ELSE T.DMBTR END AS IMPORTE_EN_ML,
        CAST(T2.BUDAT AS DATE) AS FH_CONTAB,
        T.VERTN AS NRO_CONTRATO,
        -- '----------->' AS AUXILIARES_5N,
        T.SGTXT AS TEXTO,
        T.KOART AS CL_CD,
        T.AUGDT AS COMPENSACION,
        T.ZTERM AS COND_PAGO,
        T.ZBD1T AS PLAZO_PAGO,
        T2.USNAM AS USUARIO
        FROM
            SAPABAP1.BSEG T
        LEFT JOIN SAPABAP1.BKPF T2 ON
            T.BELNR = T2.BELNR
        LEFT  JOIN SAPABAP1.KNB1 T3 ON
            T.KUNNR = T3.KUNNR
        LEFT JOIN SAPABAP1.KNA1 T4 ON 
            T.KUNNR = T4.KUNNR
        WHERE 
        T.BUKRS = 'SC10'
            AND T.HKONT IN (
        '0000121010',
        '0000121020', 
        '0000121030', 
        '0000112010', 
        '0000123010', 
        '0000121050', 
        '0000121080', 
        '0000712200', 
        '0000712400',
        '0000711201', 
        '0000711202') 
        AND T.KOART = 'D'
        AND (T.AUGDT = '00000000' OR T.AUGDT > {fh_consulta_sap}) AND T2.BUDAT <= {fh_consulta_sap}
        ORDER BY 
            T2.BELNR ASC
            """

        cursor.execute(q_5n)
        df_fbl5n = pd.read_sql_query(q_5n, conn)
        print("\t"+f"FBL5N: Se enconraron {df_fbl5n.shape[0]} partidas a fecha {fh_consulta}")
        df_fbl5n.fillna(0, inplace=True)


        return df_fbl5n



class Excel:

    
    """Manipular los excel y sumarizar saldos contables por cta ctable y cta cte de cliente"""


    def saldos_contables_y_detalle_sin_cli(self, df_fbl3n, fh_consulta):

        file_name = f'00. Composición ctas de mayor al {fh_consulta}.xlsx'
        file_name = file_name.replace("/","-")
        # Solapa Detalles ctas mayor sin clientes --> 121010 y 121050
        # Aplico formulas de columna G y J 
        print("\n"+f'Generando saldos_contables y detalle_sin_cliente')
        valores_fh_vto = []
        valores_formula_g = []
        valores_formula_j = []

        for key, values_ in df_fbl3n.iterrows():
            cta = values_[0][4:]
            values = values_[5]
            plazo_de_pago = values_[12]
            fh_doc = values_[4]  
            fh_base = values_[8]      # FH_DOC + DIAS --> Reemplazar si FH_BASE =  vacio tomar FH_DOC
            
            if fh_base == 0:
                fh_base = fh_doc    

            fh_vto = fh_base + timedelta(days=int(plazo_de_pago))        # df_fbl3n["FH_DOC"].isnull().values.any()
            fh_consulta_ = datetime.strptime(fh_consulta,'%Y/%m/%d')
            fh_consulta__ = fh_consulta_.date()
            fh_consulta__ = fh_vto
            #Calculo de anticuacion de deuda
            formula_g = fh_consulta__ - fh_vto

            # Valor col G
            if formula_g.days > 365:
                valor_g = "Partidas mayores a 365 días" 
            elif formula_g.days > 180:    
                valor_g = "Partidas entre 180-365 días"
            elif formula_g.days > 90:    
                valor_g = "Partidas entre 90-180 días"
            elif formula_g.days > 60:    
                valor_g = "Partidas entre 60-90 días"
            elif formula_g.days > 30:    
                valor_g = "Partidas entre 30-60 días"
            elif formula_g.days > 0:    
                valor_g = "Partidas entre 0-30 días"
            else:
                valor_g = "Resto partidas"

            #Valor col J
            formula_j = cta + valor_g
            valores_fh_vto.append(fh_vto)
            valores_formula_g.append(valor_g)
            valores_formula_j.append(formula_j)
        
        # Agrego valores obtenidos al df 
        df_fbl3n[fh_consulta] = valores_formula_g
        df_fbl3n['para_cruzar'] = valores_formula_j
        df_fbl3n["FH_VTO"] = valores_fh_vto

        reordenar_columnas = ["LIBRO_MAYOR","FH_CONTAB", "NRO_DOC", "CLASE", "FH_DOC","FH_VTO", fh_consulta, "IMPORTE_EN_ML", "II", "para_cruzar"]
        # Solapa saldos contables
        df_fbl3n_ = df_fbl3n.reindex(columns=reordenar_columnas)
        df_fbl3n_.to_excel(file_name ,sheet_name='saldos_contables',  index = False,)

        # Solapa detalle sin cliente para 121050 y 121080
        df_fbl3n__ = df_fbl3n_.loc[(df_fbl3n_['LIBRO_MAYOR'].isin(['0000121050','0000121080']))]
        
        # Inserto nuevo sheet
        path = file_name
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book
        df_fbl3n__.to_excel(writer, sheet_name = 'detalle_sin_cliente', index = False)
        writer.close()


        return df_fbl3n_


    def ctasmayor_con_cliente(self, df_fbl5n, fh_consulta):

        # Aplico formulas de columna G y J del archivo de trabajo
        print("\n"+f'Generando detalle_con_cliente')
        file_name = f'00. Composición ctas de mayor al {fh_consulta}.xlsx'
        file_name = file_name.replace("/","-")
        valores_fh_vto = []
        valores_formula_g = []
        valores_formula_j = []
        for key, values_ in df_fbl5n.iterrows():
            cta = values_[1][4:]
            plazo_de_pago = values_[22]
            fh_base = values_[13]                   #df_fbl5n["FH_BASE"].isnull().values.any()
            fh_vto = fh_base + timedelta(days=int(plazo_de_pago)) #Siempre tomar FH_BASE 
            fh_consulta_ = datetime.strptime(fh_consulta,'%Y/%m/%d')
            fh_consulta__ = fh_consulta_.date()
            formula_g = fh_consulta__ - fh_vto
            if formula_g.days > 365:
                valor_g = "Partidas mayores a 365 días" 
            elif formula_g.days > 180:    
                valor_g = "Partidas entre 180-365 días"
            elif formula_g.days > 90:    
                valor_g = "Partidas entre 90-180 días"
            elif formula_g.days > 60:    
                valor_g = "Partidas entre 60-90 días"
            elif formula_g.days > 30:    
                valor_g = "Partidas entre 30-60 días"
            elif formula_g.days > 0:    
                valor_g = "Partidas entre 0-30 días"
            else:
                valor_g = "Resto partidas"

            formula_j = cta + valor_g
            valores_fh_vto.append(fh_vto)
            valores_formula_g.append(valor_g)
            valores_formula_j.append(formula_j)
        
        # Agrego valores obtenidos al df 
        df_fbl5n[fh_consulta] = valores_formula_g
        df_fbl5n['para_cruzar'] = valores_formula_j
        df_fbl5n["FH_VTO"] = valores_fh_vto

        reordenar_columnas = [
            "CUENTA",
            "CLIENTE", 
            "CL", 
            "CT", 
            "VP",
            "LIBRO_MAYOR", 
            "CME", 
            "NRO_DOC",
            "REF_FACT", 
            "REFERENCIA", 
            "CLAVE_REFERENCIA", 
            "ASIGNACION", 
            "FH_DOC", 
            "FH_BASE", 
            "FH_VTO", 
            fh_consulta, 
            "IMPORTE_EN_ML", 
            "FH_CONTAB",
            "NRO_CONTRATO", 
            "TEXTO", 
            "USUARIO",
            "para_cruzar"]
        df_fbl5n_ = df_fbl5n.reindex(columns=reordenar_columnas)
        
        #Inserto detalle con cliente
        path = file_name
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book
        df_fbl5n_.to_excel(writer, sheet_name = 'detalle_con_cliente', index = False)
        writer.close()


        return df_fbl5n_


    def resumen_por_cliente(self, df_fbl3n_final, df_fbl5n_final,  fh_consulta):
        df1 = df_fbl3n_final
        df2 = df_fbl5n_final
        print("\n"+f'Generando resumen por cliente y libro mayor')
        file_name = f'00. Composición ctas de mayor al {fh_consulta}.xlsx'
        file_name = file_name.replace("/","-")

        # Si el cliente arranca con 85 marco com afiliado
        for key, values in df2.iterrows():
            cuenta = values[0]
            clasificacion = values[14]
            if cuenta.startswith("0085"):
                df2["CLIENTE"][key] = 'AFILIADO'
        df2.fillna(0, inplace = True)

        df11 = df2[['LIBRO_MAYOR','CUENTA','CLIENTE','IMPORTE_EN_ML','para_cruzar',fh_consulta]]
        
        df12 = df11.pivot_table(df11, index = ["LIBRO_MAYOR", "CUENTA","CLIENTE"], columns = [fh_consulta], aggfunc = np.sum)
        df12["total_fila"] = df12.sum(axis=1)  # FINAL
        df13 = df11.pivot_table(df11, index = ["LIBRO_MAYOR"], columns = [fh_consulta], aggfunc = np.sum)
        df13["total_fila"] = df13.sum(axis=1)   # FINAL
        # ------------
        # df3 = df2.pivot_table(df2, index = ["LIBRO_MAYOR", "CUENTA","CLIENTE"], columns = [fh_consulta], aggfunc = np.sum)
        # Total fila
        # df3["total_fila"] = df3.sum(axis=1) 
        # Total columna
        sums = df12.select_dtypes(pd.np.number).sum().rename('total_columna')
        df4 = df12.append(sums)
        # Total columna
        sums = df13.select_dtypes(pd.np.number).sum().rename('total_columna')
        df5 = df13.append(sums)

        # df5=df4.groupby(['LIBRO_MAYOR']).sum()
        
        # df5.to_excel("Resumen.xlsx",  )
        # Inserto en archivo original
        path = file_name
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book
        df12.to_excel(writer, sheet_name = 'resumen_por_cliente_bis',)
        df4.to_excel(writer, sheet_name = 'resumen_por_cliente',)
        df5.to_excel(writer, sheet_name = 'resumen_por_libro', )
        writer.close()
        print("\t"+"\t"+"\t"+"FIN DEL PROCESO")

        
       
        return 
    


# ---- Prueba -----------------------------

def proceso(fh_consulta):

    
    start = time()
    s=gmtime(start)
    print(f'Inicio proceso: {s.tm_hour - 3}:{s.tm_min} hs')

    # ------------SAP 
    partidas = Sap()
    df_fbl3n = partidas.fbl3n("PRD",fh_consulta)
    df_fbl5n = partidas.fbl5n("PRD",fh_consulta)

    # ------- Excel

    excel = Excel()
    df_fbl3n_final = excel.saldos_contables_y_detalle_sin_cli(df_fbl3n, fh_consulta)
    df_fbl5n_final = excel.ctasmayor_con_cliente(df_fbl5n, fh_consulta)
    excel.resumen_por_cliente(df_fbl3n_final, df_fbl5n_final, fh_consulta)

    tiempo = time() - start 
    print("\n"+"\t"+"\t"+f'Time: {round(tiempo/60,2)} min')
    print(f'Fin del proceso')

    return True

# --------------------------------------------------------------------------------------------------
# OK PROBADO CON TIAGO 10/2/23
# 